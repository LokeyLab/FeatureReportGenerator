import sys, os, subprocess
from datetime import date
from io import StringIO
import pandas as pd, numpy as np
import time, timeit, cython
from numba import jit, prange, njit

@jit(nopython=True, parallel=True)
def corrDist(compSig, refArray):
    def cor(u,v, centered=True): # ripped from scipy.spatial.distances
        if centered:
            umu = np.average(u)
            vmu = np.average(v)
            u = u - umu
            v = v - vmu
        uv = np.average(u*v)
        uu = np.average(np.square(u))
        vv = np.average(np.square(v))
        dist = 1 - uv / np.sqrt(uu*vv)
        return np.abs(dist)

    num_ref_sigs = refArray.shape[0]
    return_dists = np.empty(num_ref_sigs)
    
    for i in prange(num_ref_sigs):
        x = refArray[i]
        return_dists[i] = cor(compSig, x)
    
    return return_dists

@jit(nopython=True, parallel=True)
def pearsonr(compSig, refArray):
    num_ref_sigs = refArray.shape[0]
    return_corrs = np.empty(num_ref_sigs)

    for i in prange(num_ref_sigs):
        x = refArray[i,:]
        return_corrs[i] = np.corrcoef(x, compSig)[0,1]
    return return_corrs


def pairwiseCorrProcess(exp_df, ref_df, reporting_df=None, distance = True):
    if reporting_df is None:
        reporting_df = pd.DataFrame(columns=ref_df.index, index=exp_df.index)
    if distance:
        data = exp_df.apply(lambda compSig:\
            corrDist(compSig=compSig.to_numpy(), refArray=ref_df.to_numpy()).tolist(),\
            axis=1)
        # cols are the refIndex name
        data = data.apply(pd.Series)
        data.columns = list(ref_df.index)
        reporting_df = pd.DataFrame(data, columns=ref_df.index)
    else: 
        data = exp_df.apply(lambda compSig:\
            pearsonr(compSig=compSig.to_numpy(), refArray=ref_df.to_numpy()),\
            axis=1)
        data = data.apply(pd.Series)
        data.columns = list(ref_df.index)
        reporting_df = pd.DataFrame(data, columns=ref_df.index)
            
    return reporting_df

# def createSubDf(i, distDf, pearsonDf): # instead of concatinating why not create a new df?
#     '''Create a sheet for a xlsx workbook'''
#     pageName = "._.".join(map(str, i))
#     combDf = pd.DataFrame({
#         'Pearson_R': pearsonDf[i],
#         'Corr_Dist': distDf[i]
#     })
#     return combDf, pageName

# def createXLSheet(distanceReport: pd.DataFrame, pearsonReport: pd.DataFrame, outName: str):
#     assert list(distanceReport.columns) == list(pearsonReport.columns)

#     print("Creating Excel file...", file=sys.stderr)
#     with pd.ExcelWriter(outName) as f:
#         for i in distanceReport.columns: #this actually takes so long to process, it might be better to cache some files
#             combDf, pageName = createSubDf(i, distDf=distanceReport, pearsonDf=pearsonReport)
#             combDf.to_excel(f, sheet_name=pageName)

#     print("...Completed!", file=sys.stderr)
#     return

import multiprocessing as mp
from openpyxl import load_workbook, Workbook
class CreateXLSheetMultithread:
    def __init__(self, cwd, distanceReport: pd.DataFrame, pearsonReport: pd.DataFrame, outName: str, threads: int = None, verbose = False, singleIndex = True):
        self.cwd = cwd
        self.distDf = distanceReport
        self.pearsonDf = pearsonReport
        self.threads = threads
        self.outName = outName
        self.verbose = verbose
        self.singleIndex = singleIndex

        self.folderPath = os.path.join(self.cwd, '.temp/')
    
    def __cleanTempFiles(self):
        folderPath = self.folderPath

        if os.path.exists(folderPath):
            for file in os.listdir(folderPath):
                filePath = os.path.join(folderPath, file)

                if os.path.isfile(filePath):
                    os.remove(filePath)
            os.rmdir(folderPath)
        return
    
    def __chunkToExcel(self, chunk: dict, outName: str):
        with pd.ExcelWriter(outName, engine='openpyxl') as writer:
            for sheetName, df in chunk.items():

                if self.verbose:
                    print(f'...Working on: {sheetName}...', file=sys.stderr)

                df.to_excel(writer, sheet_name=sheetName, index=True)
    
    def processhandler(self, startIndex, endIndex, data: list, notebookDir):
        chunk = {d[1]:d[0] for d in data[startIndex:endIndex]}
        nbName = f'temp_{startIndex}.xlsx'
        name = os.path.join(self.cwd, '.temp/', nbName)

        self.__chunkToExcel(chunk=chunk, outName=name)
    
    def __createSubDf(self, i, distDf, pearsonDf):
        if self.singleIndex:
            pageName = i
        else:
            pageName = "._.".join(map(str, i))

        combDf = pd.DataFrame({
            'Pearson_R': pearsonDf[i],
            'Corr_Dist': distDf[i]
        })

        return combDf, pageName

    def parallelWrite(self):
        distanceReport = self.distDf
        pearsonReport = self.pearsonDf
        cwd = self.cwd

        try:
            os.mkdir('.temp/')
        except:
            self.__cleanTempFiles()
            os.mkdir('.temp/')
        
        sheets = [self.__createSubDf(i, distDf=distanceReport, pearsonDf=pearsonReport) for i in distanceReport.columns]
        numCPUs = (mp.cpu_count() - 2) if self.threads is None else (self.threads - 1) 
        chunkSize = len(sheets) // numCPUs

        if self.verbose:
            print(f'Number of thread used in the program: {numCPUs} with 1 or 2 threads left as a spare', file=sys.stderr)
            print(f'Partition Size: {chunkSize}', file=sys.stderr)

            print('Beginning Multithreading process...', file=sys.stderr)

        processes = []
        for start in range(0, len(sheets), chunkSize):
            end = start + chunkSize if start + chunkSize <= len(sheets) else len(sheets)

            process = mp.Process(target=self.processhandler, args=(start, end, sheets, self.cwd))
            processes.append(process)
            process.start()
        
        for process in processes:
            process.join()

        if self.verbose:
            print("...Multithreading process finished!", file=sys.stderr)
        
    def combNotebooks(self):
        tempNotebooks = []
        if os.path.exists(path=self.folderPath): # sanity check
            tempNotebooks = [os.path.join(self.folderPath, file) for file in os.listdir(self.folderPath)]
        else:
            print("\n.temp/ folder not found", file=sys.stderr)
            exit(-1)
        
        mergedWb = Workbook()
        mergedWb.remove(mergedWb.active)

        if self.verbose:
            print("Merging all partitioned notebooks", file=sys.stderr)

        for i, notebook in enumerate(tempNotebooks):
            if self.verbose:
                print(f'Working on notebook {i}', file=sys.stderr)

            workbook = load_workbook(filename=notebook, read_only=True)

            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                newWs = mergedWb.create_sheet(title = sheet)

                for row in ws.iter_rows():
                    newRow = [cell.value for cell in row]
                    newWs.append(newRow)

        mergedWb.save(self.outName)

        if self.verbose:
            print("Merging done!", file=sys.stderr)
        self.__cleanTempFiles()

class CommandLine:
    def __init__(self, inOpts = None):
        import argparse

        # parser
        self.parser = argparse.ArgumentParser(
            description='Generates a Feature Report (xlsx file) that shows the reference compounds and its Correlation Distance and its Pearson R score',
            add_help=True,
            prefix_chars='-',
            usage='python3 %(prog)s -e <experimental wells> -r <reference wells> -o <output name> [-options] [argument]'
        )

        # arguments
        self.parser.add_argument('-e', '--experimental', type=str, nargs='?', action='store', help='file input for experimental wells (.csv file accepted only)', required=True)
        self.parser.add_argument('-r', '--reference', type=str, nargs='?', action='store', help='file input for experimental wells (.csv files accepted only)', required=True)
        self.parser.add_argument('-o', '--out', type=str, nargs='?', action='store', help='name for output file in xlsx format (make sure it ends in .xlsx)', required=True)
        self.parser.add_argument('-t', '--threads', default=8, type=int, action='store', nargs='?', help='Number of threads to use for feature report writing (default: 8)')
        self.parser.add_argument('-i', '--index', default=[0,1,2,3], type=int, action='store', nargs='+', help='Specifies which columns of the input files are the index columns (default: 0 1 2 3)')
        self.parser.add_argument('-w', '--wells', default=None, type=str,action='store', nargs='+', help='Specifies which wells specifically to process from the experiment.')
        self.parser.add_argument('-v', '--verbose', default=False, action='store_true', help='Enables verbose output to stderr (default: False) Note: I recommend having this flag enabled when testing or building with this program')

        # arg parsing
        if inOpts is None:
            self.args = self.parser.parse_args()
        else:
            self.args = self.parser.parse_args(inOpts)

def pairwiseCorrDistProcess(expDf: pd.DataFrame, refDf: pd.DataFrame):
    return pairwiseCorrProcess(expDf, refDf, distance=True)

def pairwisePearsonProcess(expDf: pd.DataFrame, refDf: pd.DataFrame):
    return pairwiseCorrProcess(expDf, refDf, distance=False)

def main(inOpts = None):
    from time import sleep, time
    # allDF = pd.read_csv("data/AllReferenceExp_20191018_commonFeaturesOrder.csv",index_col=[0,1,2,3])
    # testDF = pd.read_csv("data/SP0142_20171024_HeLa_10x_0_CP_histdiff_Concatenated.csv",index_col=[0,1,2,3])
    # testDF = testDF.reindex(columns = allDF.columns)
    # # print(allDF.to_numpy().shape)
    # # cols= 1) (optional) sheet name 2) relationship to ref 3) dist 4) R
    # testSig_comp = testDF.index[0]
    # testSig = testDF.loc[testSig_comp]
    #print(testSig.to_numpy().shape[0])
    # print(len(allDF.iloc[0].to_numpy()))

    ## Above are test files ##
    ###### Main program below this commented line ######
    cl = CommandLine(inOpts=inOpts)

    singleIndex =  False if len(cl.args.index) > 1 else True

    expDf = pd.read_csv(cl.args.experimental, sep=',', index_col=cl.args.index)
    if cl.args.wells is not None:
        # print(cl.args.wells,file=sys.stderr)
        expDf.drop(index=expDf.index.difference(cl.args.wells),inplace=True) #.copy()
        
    refDf = pd.read_csv(cl.args.reference, sep=',', index_col=cl.args.index)
    outName = cl.args.out
    threads = cl.args.threads
    verbose = cl.args.verbose

    start = time()

    # distance = pairwiseCorrProcess(exp_df=expDf, ref_df=refDf, distance=True)
    # distance = distance.transpose()

    # pearson = pairwiseCorrProcess(exp_df=expDf, ref_df=refDf, distance=False)
    # pearson = pearson.transpose()

    # try to make the calculations parallelized
    with mp.Pool(processes=2) as pool:

        distanceAsync = pool.apply_async(pairwiseCorrDistProcess, args=(expDf, refDf))
        pearsonAsync = pool.apply_async(pairwisePearsonProcess, args=(expDf, refDf))

        distance = distanceAsync.get().transpose()
        pearson = pearsonAsync.get().transpose()
    xlsxgen = CreateXLSheetMultithread(cwd=os.getcwd(), distanceReport=distance, pearsonReport=pearson, outName=outName, threads=threads, verbose=verbose, singleIndex=singleIndex)
    xlsxgen.parallelWrite()
    sleep(1) # wait for files to flush to disk
    xlsxgen.combNotebooks()

    if verbose:
        runTime = time()-start
        print(f'done! Completed in {int(runTime // 60)} min(s) {int(runTime % 60)} sec(s)', file=sys.stderr)

if __name__ =='__main__':
    main()